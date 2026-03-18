#!/usr/bin/env python3
"""
sync_listings.py - Sync property listings from Google Sheets to PostgreSQL.
Reads "Listing CPA Online" spreadsheet, compares with DB, inserts/updates as needed.

Usage:
  python sync_listings.py                  # Full sync
  python sync_listings.py --dry-run        # Preview changes without writing
  python sync_listings.py --sheet 1        # Sync only Sheet 1 (JUAL & SEWA)
  python sync_listings.py --sheet 2        # Sync only Sheet 2 (DEVELOPER)
  python sync_listings.py --sync-images    # Also trigger image sync for new listings
"""

import os
import sys
import json
import re
import time
import logging
import subprocess
import argparse
import tempfile
from datetime import datetime

import io
import openpyxl
import requests
from color_coding import fetch_cell_formatting, get_color_status, fetch_cell_formatting_xlsx

# ─── Configuration ───────────────────────────────────────────────────────────

SPREADSHEET_ID = "10ll9ZxNfkD6PgAmjvEed_oj5Qg6AeZP6"
TOKEN_FILE = "/var/www/livininbintaro/api/.google_token.json"
LOG_FILE = "/var/www/livininbintaro/api/sync.log"
SYNC_IMAGES_SCRIPT = "/var/www/livininbintaro/api/sync_images.py"

SHEET1_NAME = "LISTING JUAL & SEWA"  # CPA listings
SHEET2_NAME = "LISTING DEVELOPER"     # Developer listings

# Status keywords in Keterangan column
SOLD_KEYWORDS = ["terjual", "sold", "laku", "deal"]
CANCELLED_KEYWORDS = ["cancel", "batal", "void"]

# SQL batch size (how many statements to run at once)
SQL_BATCH_SIZE = 500

# ─── Logging Setup ───────────────────────────────────────────────────────────

def setup_logging(dry_run=False):
    """Configure logging to both file and console."""
    prefix = "[DRY-RUN] " if dry_run else ""
    formatter = logging.Formatter(
        f'%(asctime)s {prefix}%(levelname)s: %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    logger = logging.getLogger("sync_listings")
    logger.setLevel(logging.DEBUG)

    # File handler
    fh = logging.FileHandler(LOG_FILE, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(formatter)
    logger.addHandler(fh)

    # Console handler
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    ch.setFormatter(formatter)
    logger.addHandler(ch)

    return logger

# ─── Google Sheets API ───────────────────────────────────────────────────────

# ─── File type detection & xlsx support ────────────────────────────────────

_file_mime = None   # cached mimeType
_xlsx_wb   = None   # cached openpyxl workbook

def detect_file_type(token):
    """Return mimeType of the spreadsheet file via Drive API."""
    global _file_mime
    if _file_mime:
        return _file_mime
    url = f"https://www.googleapis.com/drive/v3/files/{SPREADSHEET_ID}?fields=mimeType,name"
    resp = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=15)
    resp.raise_for_status()
    data = resp.json()
    _file_mime = data.get("mimeType", "")
    return _file_mime

def load_xlsx_workbook(token):
    """Download Excel file from Drive and return openpyxl workbook (cached)."""
    global _xlsx_wb
    if _xlsx_wb:
        return _xlsx_wb
    url = f"https://www.googleapis.com/drive/v3/files/{SPREADSHEET_ID}?alt=media"
    resp = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=120, stream=True)
    resp.raise_for_status()
    _xlsx_wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True, read_only=False)
    return _xlsx_wb

def is_xlsx(mime):
    """True if file is Excel format (not native Google Sheets)."""
    return "spreadsheet" not in mime or "google" not in mime

def fetch_sheet_data_xlsx(wb, sheet_name):
    """Fetch all data rows from an openpyxl sheet. Returns list of rows (list of str)."""
    ws = None
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        for name in wb.sheetnames:
            if name.lower() == sheet_name.lower():
                ws = wb[name]
                break
    if ws is None:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    rows = []
    for row in ws.iter_rows(values_only=True):
        cells = []
        for v in row:
            if v is None:
                cells.append("")
            elif isinstance(v, bool):
                cells.append(str(v))
            else:
                cells.append(str(v).strip())
        # Trim trailing empty
        while cells and cells[-1] == "":
            cells.pop()
        rows.append(cells)
    # Trim trailing empty rows
    while rows and not any(rows[-1]):
        rows.pop()
    return rows

# ─── Google Sheets/Drive API ─────────────────────────────────────────────────

def get_access_token():
    """Get a fresh access token using the refresh token from token file."""
    # Load client credentials
    with open("/root/.secrets/livininbintaro-google.json") as f:
        client_creds = json.load(f)["installed"]
    
    # Load refresh token
    with open(TOKEN_FILE) as f:
        token_data = json.load(f)
    
    resp = requests.post("https://oauth2.googleapis.com/token", data={
        "client_id": client_creds["client_id"],
        "client_secret": client_creds["client_secret"],
        "refresh_token": token_data["refresh_token"],
        "grant_type": "refresh_token",
    }, timeout=15)
    resp.raise_for_status()
    return resp.json()["access_token"]
def fetch_sheet_data(token, sheet_name, range_suffix=""):
    """Fetch all data from a sheet. Returns list of rows."""
    encoded_name = sheet_name.replace("&", "%26")
    range_str = f"{encoded_name}!A1:AQ" if not range_suffix else f"{encoded_name}!{range_suffix}"

    url = f"https://sheets.googleapis.com/v4/spreadsheets/{SPREADSHEET_ID}/values/{range_str}"
    headers = {"Authorization": f"Bearer {token}"}
    params = {"valueRenderOption": "FORMATTED_VALUE"}

    resp = requests.get(url, headers=headers, params=params, timeout=120)
    resp.raise_for_status()
    data = resp.json()
    return data.get("values", [])

# ─── Database Operations ─────────────────────────────────────────────────────

def run_sql(sql, logger=None):
    """Run SQL via podman exec psql, return stdout."""
    cmd = [
        "docker", "exec", "-i", "livininbintaro-db",
        "psql", "-U", "livin", "-d", "livininbintaro",
        "-t", "-A", "-c", sql
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
    if result.returncode != 0:
        if logger:
            logger.error(f"SQL ERROR: {result.stderr.strip()}")
        return ""
    return result.stdout.strip()


def run_sql_batch(sql_statements, logger=None):
    """Run multiple SQL statements in a single transaction via psql pipe.
    Much faster than individual podman exec calls."""
    if not sql_statements:
        return True

    # Wrap in transaction with SAVEPOINT for error recovery
    full_sql = "BEGIN;\n"
    full_sql += "\n".join(sql_statements)
    full_sql += "\nCOMMIT;\n"

    # Write to temp file and pipe through psql
    with tempfile.NamedTemporaryFile(mode='w', suffix='.sql', delete=False) as f:
        f.write(full_sql)
        tmp_path = f.name

    try:
        # Copy file into container and execute
        subprocess.run(
            ["docker", "cp", tmp_path, "livininbintaro-db:/tmp/batch.sql"],
            capture_output=True, text=True, timeout=30
        )

        result = subprocess.run(
            ["docker", "exec", "-i", "livininbintaro-db",
             "psql", "-U", "livin", "-d", "livininbintaro",
             "-f", "/tmp/batch.sql"],
            capture_output=True, text=True, timeout=300
        )

        if result.returncode != 0:
            if logger:
                # Extract first error line for clarity
                error_lines = result.stderr.strip().split('\n')
                for line in error_lines:
                    if "ERROR" in line:
                        logger.error(f"SQL BATCH ERROR: {line[:200]}")
                        break
                else:
                    logger.error(f"SQL BATCH ERROR: {result.stderr.strip()[:500]}")
            return False

        # Check for errors in output
        if "ERROR" in (result.stderr or ""):
            if logger:
                error_lines = result.stderr.strip().split('\n')
                for line in error_lines:
                    if "ERROR" in line and "duplicate key" not in line.lower():
                        logger.error(f"SQL BATCH ERROR: {line[:200]}")
                        break
            return False

        return True
    finally:
        os.unlink(tmp_path)
        subprocess.run(
            ["docker", "exec", "-i", "livininbintaro-db", "rm", "-f", "/tmp/batch.sql"],
            capture_output=True, text=True, timeout=10
        )


def get_existing_listings(logger):
    """Load all existing listings from DB, keyed by normalized CPA code."""
    sql = (
        "SELECT id, cpa_code, title, slug, description, property_type, listing_type, "
        "price, price_label, area_location, cluster, sektor, address, block, block_no, "
        "bedrooms, bedrooms_extra, bathrooms, bathrooms_extra, land_area, building_area, "
        "floors, certificate, electricity, cpa_expiry, drive_folder_id, drive_link, "
        "summary_client, summary_ma, is_hot, is_active "
        "FROM listings ORDER BY id"
    )
    raw = run_sql(sql, logger)
    if not raw:
        return {}

    listings = {}
    for line in raw.split("\n"):
        line = line.strip()
        if not line:
            continue
        parts = line.split("|")
        if len(parts) < 31:
            continue

        cpa_code_raw = parts[1]
        cpa_key = normalize_cpa(cpa_code_raw)

        listings[cpa_key] = {
            "id": int(parts[0]),
            "cpa_code": parts[1],
            "title": parts[2],
            "slug": parts[3],
            "description": parts[4],
            "property_type": parts[5],
            "listing_type": parts[6],
            "price": int(parts[7]) if parts[7] else None,
            "price_label": parts[8],
            "area_location": parts[9],
            "cluster": parts[10],
            "sektor": parts[11],
            "address": parts[12],
            "block": parts[13],
            "block_no": parts[14],
            "bedrooms": int(parts[15]) if parts[15] else None,
            "bedrooms_extra": int(parts[16]) if parts[16] else None,
            "bathrooms": int(parts[17]) if parts[17] else None,
            "bathrooms_extra": int(parts[18]) if parts[18] else None,
            "land_area": int(parts[19]) if parts[19] else None,
            "building_area": int(parts[20]) if parts[20] else None,
            "floors": int(float(parts[21])) if parts[21] else None,
            "certificate": parts[22],
            "electricity": normalize_electricity(parts[23]),
            "cpa_expiry": parts[24],
            "drive_folder_id": parts[25],
            "drive_link": parts[26],
            "summary_client": parts[27],
            "summary_ma": parts[28],
            "is_hot": parts[29] == "t",
            "is_active": parts[30] == "t",
        }

    logger.info(f"Loaded {len(listings)} existing listings from DB")
    return listings

# ─── Data Parsing & Normalization ────────────────────────────────────────────

def normalize_cpa(cpa_code):
    """Normalize CPA code: strip, collapse internal whitespace."""
    if not cpa_code:
        return ""
    return re.sub(r'\s+', ' ', cpa_code.strip())


def normalize_electricity(val):
    """Normalize electricity value: strip trailing .0"""
    if not val or not val.strip():
        return None
    val = val.strip()
    # Remove trailing .0 (from float import)
    val = re.sub(r'\.0$', '', val)
    return val if val else None


def parse_price(price_str):
    """Parse price string like 'Rp5.800.000.000' to integer."""
    if not price_str or not price_str.strip():
        return None
    cleaned = re.sub(r'[Rp\s.]', '', price_str.strip())
    cleaned = re.sub(r'[^\d]', '', cleaned)
    if cleaned:
        try:
            return int(cleaned)
        except ValueError:
            return None
    return None


def parse_int(val):
    """Parse integer from string, handling floats like '2.5' → 2."""
    if not val or not val.strip():
        return None
    val = val.strip().replace(',', '.')
    try:
        f = float(val)
        return int(f)
    except (ValueError, TypeError):
        return None


def map_property_type(val):
    """Map spreadsheet property type to DB enum."""
    if not val:
        return None
    val_lower = val.strip().lower()
    mapping = {
        "rumah": "rumah",
        "apartemen": "apartemen",
        "apartment": "apartemen",
        "kavling": "kavling",
        "tanah": "tanah",
        "ruko": "ruko",
        "rukan": "ruko",
        "kios": "ruko",
    }
    return mapping.get(val_lower, val_lower)


def map_listing_type(val):
    """Map J/S to listing type."""
    if not val:
        return None
    val = val.strip().upper()
    if val == "J":
        return "dijual"
    elif val == "S":
        return "disewa"
    return None


def detect_status(keterangan):
    """Detect sold/cancelled status from Keterangan text.
    Only triggers if the keyword is the MAIN content (not just mentioned in passing)."""
    if not keterangan:
        return True, None
    keterangan_lower = keterangan.strip().lower()

    # Only mark as sold/cancelled if the text is primarily about that status
    # (short text or starts with the keyword)
    for kw in SOLD_KEYWORDS:
        if keterangan_lower == kw or keterangan_lower.startswith(kw):
            return False, f"sold:{kw}"
    for kw in CANCELLED_KEYWORDS:
        if keterangan_lower == kw or keterangan_lower.startswith(kw):
            return False, f"cancelled:{kw}"
    return True, None


def slugify(text):
    """Generate URL slug from text."""
    text = text.lower().strip()
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[\s_-]+', '-', text)
    return text


def apply_color_status(listing, row_index, sheet_colors):
    """Apply color-based status to listing from cell formatting.
    
    Args:
        listing: listing dict with parsed data
        row_index: 0-based row number in sheet
        sheet_colors: dict mapping row_index -> color dict
    
    Returns:
        updated listing dict
    """
    if row_index not in sheet_colors:
        return listing
    
    bg_color = sheet_colors[row_index]
    is_active, is_hot, hot_reason = get_color_status(bg_color)
    
    # Apply color-detected status (takes precedence over keyword detection)
    listing["is_active"] = is_active
    listing["is_hot"] = is_hot
    listing["hot_reason"] = hot_reason
    
    return listing


def escape_sql(val):
    """Escape single quotes for SQL."""
    if val is None:
        return "NULL"
    if isinstance(val, bool):
        return "true" if val else "false"
    if isinstance(val, (int, float)):
        return str(int(val)) if isinstance(val, int) else str(val)
    s = str(val).replace("'", "''")
    return f"'{s}'"

# ─── Row Parsing ─────────────────────────────────────────────────────────────

def get_cell(row, idx, default=""):
    """Safely get a cell value from a row."""
    if idx < len(row):
        val = row[idx]
        if val and val.strip():
            return val.strip()
    return default


def parse_sheet1_row(row):
    """Parse a row from Sheet 1 (LISTING JUAL & SEWA) into a listing dict."""
    cpa_raw = get_cell(row, 1)
    if not cpa_raw:
        return None

    cpa_key = normalize_cpa(cpa_raw)
    if not cpa_key.startswith("CPA"):
        return None

    alamat = get_cell(row, 4)
    keterangan = get_cell(row, 32)
    rangkuman_client = get_cell(row, 34)
    rangkuman_ma = get_cell(row, 35)
    harga_ringkas = get_cell(row, 37)
    alamat_db = get_cell(row, 38)

    is_active, status_reason = detect_status(keterangan)
    description = rangkuman_client or keterangan or ""

    return {
        "cpa_code": cpa_raw,
        "cpa_key": cpa_key,
        "title": alamat,
        "cluster": alamat,
        "area_location": get_cell(row, 8) or None,
        "sektor": get_cell(row, 9) or None,
        "property_type": map_property_type(get_cell(row, 10)),
        "listing_type": map_listing_type(get_cell(row, 11)),
        "land_area": parse_int(get_cell(row, 12)),
        "building_area": parse_int(get_cell(row, 13)),
        "price": parse_price(get_cell(row, 14)),
        "bedrooms": parse_int(get_cell(row, 16)),
        "bedrooms_extra": parse_int(get_cell(row, 17)),
        "bathrooms": parse_int(get_cell(row, 18)),
        "bathrooms_extra": parse_int(get_cell(row, 19)),
        "floors": parse_int(get_cell(row, 21)),
        "electricity": normalize_electricity(get_cell(row, 25)),
        "certificate": get_cell(row, 28) or None,
        "block": get_cell(row, 5) or None,
        "block_no": get_cell(row, 6) or None,
        "description": description,
        "summary_client": rangkuman_client or None,
        "summary_ma": rangkuman_ma or None,
        "price_label": harga_ringkas or None,
        "address": alamat_db or None,
        "is_active": is_active,
        "status_reason": status_reason,
    }


def parse_sheet2_row(row):
    """Parse a row from Sheet 2 (LISTING DEVELOPER) into a listing dict."""
    cpa_raw = get_cell(row, 0)
    if not cpa_raw:
        return None

    cpa_key = normalize_cpa(cpa_raw)
    if not cpa_key.startswith("DVLP"):
        return None

    alamat = get_cell(row, 3)
    keterangan = get_cell(row, 29)
    rangkuman_client = get_cell(row, 31)
    rangkuman_ma = get_cell(row, 32)
    harga_ringkas = get_cell(row, 34)
    alamat_db = get_cell(row, 35)

    is_active, status_reason = detect_status(keterangan)
    description = rangkuman_client or keterangan or ""

    return {
        "cpa_code": cpa_raw,
        "cpa_key": cpa_key,
        "title": alamat,
        "cluster": alamat,
        "area_location": None,
        "sektor": None,
        "property_type": map_property_type(get_cell(row, 28)),
        "listing_type": map_listing_type(get_cell(row, 25)),
        "land_area": parse_int(get_cell(row, 6)),
        "building_area": parse_int(get_cell(row, 7)),
        "price": parse_price(get_cell(row, 8)),
        "bedrooms": parse_int(get_cell(row, 11)),
        "bedrooms_extra": parse_int(get_cell(row, 12)),
        "bathrooms": parse_int(get_cell(row, 13)),
        "bathrooms_extra": parse_int(get_cell(row, 14)),
        "floors": parse_int(get_cell(row, 16)),
        "electricity": normalize_electricity(get_cell(row, 20)),
        "certificate": get_cell(row, 23) or None,
        "block": get_cell(row, 4) or None,
        "block_no": get_cell(row, 5) or None,
        "description": description,
        "summary_client": rangkuman_client or None,
        "summary_ma": rangkuman_ma or None,
        "price_label": harga_ringkas or None,
        "address": alamat_db or None,
        "is_active": is_active,
        "status_reason": status_reason,
    }

# ─── Sync Logic ──────────────────────────────────────────────────────────────

# Fields to compare for detecting updates
COMPARE_FIELDS = [
    "title", "property_type", "listing_type", "price", "price_label",
    "area_location", "cluster", "sektor", "address", "block", "block_no",
    "bedrooms", "bedrooms_extra", "bathrooms", "bathrooms_extra",
    "land_area", "building_area", "floors", "certificate", "electricity",
    "description", "summary_client", "summary_ma", "is_active",
]


def normalize_for_compare(val):
    """Normalize a value for comparison."""
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return int(val) if val == int(val) else val
    if isinstance(val, str):
        val = val.strip()
        # Normalize trailing .0 in numeric strings
        val = re.sub(r'\.0$', '', val)
        return val if val else None
    return val


def find_changes(existing, new_data):
    """Compare existing DB record with new spreadsheet data. Returns dict of changed fields."""
    changes = {}
    for field in COMPARE_FIELDS:
        old_val = normalize_for_compare(existing.get(field))
        new_val = normalize_for_compare(new_data.get(field))

        # Treat None/empty as equivalent
        if not old_val and not new_val:
            continue
        if old_val is None and new_val is None:
            continue

        # Compare
        if old_val != new_val:
            # Don't overwrite existing data with empty new data (preserve existing)
            if not new_val and old_val:
                continue
            changes[field] = {"old": old_val, "new": new_data.get(field)}

    return changes


def generate_unique_slug(title, existing_slugs):
    """Generate a unique slug, appending counter if needed."""
    base_slug = slugify(title) if title else "listing"
    if not base_slug:
        base_slug = "listing"

    slug = base_slug
    counter = 1
    while slug in existing_slugs:
        counter += 1
        slug = f"{base_slug}-{counter}"

    existing_slugs.add(slug)
    return slug


def build_insert_sql(listing_data, slug):
    """Build INSERT SQL for a new listing with ON CONFLICT handling."""
    fields = [
        "cpa_code", "title", "slug", "cluster", "description", "property_type",
        "listing_type", "price", "price_label", "area_location", "sektor",
        "address", "block", "block_no", "bedrooms", "bedrooms_extra",
        "bathrooms", "bathrooms_extra", "land_area", "building_area",
        "floors", "certificate", "electricity", "summary_client", "summary_ma",
        "is_active", "created_at", "updated_at"
    ]

    values = [
        escape_sql(listing_data["cpa_code"]),
        escape_sql(listing_data.get("title")),
        escape_sql(slug),
        escape_sql(listing_data.get("cluster")),
        escape_sql(listing_data.get("description")),
        escape_sql(listing_data.get("property_type")),
        escape_sql(listing_data.get("listing_type")),
        escape_sql(listing_data.get("price")),
        escape_sql(listing_data.get("price_label")),
        escape_sql(listing_data.get("area_location")),
        escape_sql(listing_data.get("sektor")),
        escape_sql(listing_data.get("address")),
        escape_sql(listing_data.get("block")),
        escape_sql(listing_data.get("block_no")),
        escape_sql(listing_data.get("bedrooms")),
        escape_sql(listing_data.get("bedrooms_extra")),
        escape_sql(listing_data.get("bathrooms")),
        escape_sql(listing_data.get("bathrooms_extra")),
        escape_sql(listing_data.get("land_area")),
        escape_sql(listing_data.get("building_area")),
        escape_sql(listing_data.get("floors")),
        escape_sql(listing_data.get("certificate")),
        escape_sql(listing_data.get("electricity")),
        escape_sql(listing_data.get("summary_client")),
        escape_sql(listing_data.get("summary_ma")),
        escape_sql(listing_data.get("is_active", True)),
        "NOW()",
        "NOW()",
    ]

    # Use ON CONFLICT DO NOTHING to avoid duplicate key errors
    # If slug already exists, skip insert (should be caught by generate_unique_slug, but just in case)
    return f"INSERT INTO listings ({', '.join(fields)}) VALUES ({', '.join(values)}) ON CONFLICT (slug) DO NOTHING;"


def build_update_sql(listing_id, changes):
    """Build UPDATE SQL for changed fields."""
    set_parts = []
    for field, vals in changes.items():
        set_parts.append(f"{field} = {escape_sql(vals['new'])}")
    set_parts.append("updated_at = NOW()")

    return f"UPDATE listings SET {', '.join(set_parts)} WHERE id = {listing_id};"

# ─── Main Sync Function ─────────────────────────────────────────────────────

def sync(args):
    """Main sync function."""
    dry_run = args.dry_run
    sync_images = args.sync_images
    sheet_filter = args.sheet

    logger = setup_logging(dry_run)
    logger.info("=" * 70)
    logger.info(f"SYNC STARTED at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info(f"Mode: {'DRY-RUN' if dry_run else 'LIVE'}")
    logger.info(f"Sheets: {'Both' if sheet_filter == 0 else f'Sheet {sheet_filter} only'}")
    logger.info("=" * 70)

    stats = {
        "sheet1_rows": 0, "sheet2_rows": 0,
        "new": 0, "updated": 0, "deactivated": 0,
        "unchanged": 0, "skipped": 0, "errors": 0,
        "new_cpa_codes": [],
    }
    batch_count = 0
    executed_batch_count = 0  # init here so summary block always has it

    try:
        # 1. Get OAuth token
        logger.info("Getting OAuth access token...")
        token = get_access_token()
        logger.info("Token acquired successfully")

        # 1b. Detect file type (Google Sheets vs Excel)
        logger.info("Detecting spreadsheet file type...")
        mime = detect_file_type(token)
        use_xlsx = is_xlsx(mime)
        logger.info(f"File type: {mime} → {'Excel/xlsx mode' if use_xlsx else 'Google Sheets API mode'}")

        wb = None
        if use_xlsx:
            logger.info("Downloading xlsx workbook from Drive...")
            wb = load_xlsx_workbook(token)
            logger.info(f"Workbook loaded. Sheets: {wb.sheetnames}")

        # 2. Load existing data from DB
        logger.info("Loading existing listings from database...")
        existing = get_existing_listings(logger)

        existing_slugs = set()
        for rec in existing.values():
            if rec["slug"]:
                existing_slugs.add(rec["slug"])

        # 3. Fetch spreadsheet data and cell formatting
        all_sheet_data = []

        # Fetch color formatting for both sheets
        logger.info("Fetching cell formatting (colors)...")
        sheet1_colors = {}
        sheet2_colors = {}
        try:
            if use_xlsx:
                sheet1_colors_list = fetch_cell_formatting_xlsx(wb, SHEET1_NAME, col_idx=2)
            else:
                sheet1_colors_list = fetch_cell_formatting(token, SHEET1_NAME, "B:B")
            sheet1_colors = {row["row"]: row["color"] for row in sheet1_colors_list if row["row"] > 0}
            logger.info(f"Sheet 1: fetched {len(sheet1_colors)} color entries")
        except Exception as e:
            logger.warning(f"Could not fetch Sheet 1 colors: {e}")

        try:
            if use_xlsx:
                sheet2_colors_list = fetch_cell_formatting_xlsx(wb, SHEET2_NAME, col_idx=1)
            else:
                sheet2_colors_list = fetch_cell_formatting(token, SHEET2_NAME, "A:A")
            sheet2_colors = {row["row"]: row["color"] for row in sheet2_colors_list if row["row"] > 0}
            logger.info(f"Sheet 2: fetched {len(sheet2_colors)} color entries")
        except Exception as e:
            logger.warning(f"Could not fetch Sheet 2 colors: {e}")

        if sheet_filter in (0, 1):
            logger.info(f"Fetching Sheet 1 ({SHEET1_NAME})...")
            if use_xlsx:
                sheet1_rows = fetch_sheet_data_xlsx(wb, SHEET1_NAME)
            else:
                sheet1_rows = fetch_sheet_data(token, SHEET1_NAME)
            stats["sheet1_rows"] = len(sheet1_rows) - 1
            logger.info(f"Sheet 1: {stats['sheet1_rows']} data rows (+ header)")

            for i, row in enumerate(sheet1_rows[1:], start=2):
                parsed = parse_sheet1_row(row)
                if parsed:
                    parsed["_sheet"] = 1
                    parsed["_row"] = i
                    parsed = apply_color_status(parsed, i-1, sheet1_colors)
                    all_sheet_data.append(parsed)

        if sheet_filter in (0, 2):
            logger.info(f"Fetching Sheet 2 ({SHEET2_NAME})...")
            if use_xlsx:
                sheet2_rows = fetch_sheet_data_xlsx(wb, SHEET2_NAME)
            else:
                sheet2_rows = fetch_sheet_data(token, SHEET2_NAME)
            stats["sheet2_rows"] = len(sheet2_rows) - 1
            logger.info(f"Sheet 2: {stats['sheet2_rows']} data rows (+ header)")

            for i, row in enumerate(sheet2_rows[1:], start=2):
                parsed = parse_sheet2_row(row)
                if parsed:
                    parsed["_sheet"] = 2
                    parsed["_row"] = i
                    parsed = apply_color_status(parsed, i-1, sheet2_colors)
                    all_sheet_data.append(parsed)

        logger.info(f"Total parsed listings from sheets: {len(all_sheet_data)}")

        # 4. Process each listing - collect SQL statements
        sql_batch = []

        for idx, listing in enumerate(all_sheet_data):
            cpa_key = listing["cpa_key"]

            if not listing.get("title"):
                stats["skipped"] += 1
                continue

            # Skip sold/cancelled listings that already exist in DB with same status
            if not listing.get("is_active", True) and cpa_key in existing:
                db_rec = existing[cpa_key]
                if not db_rec.get("is_active", True):
                    stats["skipped"] += 1
                    continue

            try:
                if cpa_key in existing:
                    db_rec = existing[cpa_key]
                    changes = find_changes(db_rec, listing)

                    if changes:
                        changed_fields = list(changes.keys())
                        # Only log significant changes (not just summary enrichment)
                        significant = [f for f in changed_fields if f not in ("summary_client", "summary_ma")]
                        if significant:
                            logger.info(f"UPDATE {cpa_key} (id={db_rec['id']}): {changed_fields}")
                        else:
                            logger.debug(f"ENRICH {cpa_key} (id={db_rec['id']}): {changed_fields}")

                        if not dry_run:
                            sql = build_update_sql(db_rec["id"], changes)
                            sql_batch.append(sql)
                        
                        stats["updated"] += 1
                        if "is_active" in changes and not changes["is_active"]["new"]:
                            stats["deactivated"] += 1
                            logger.info(f"  DEACTIVATED: {cpa_key} ({listing.get('status_reason', 'unknown')})")
                    else:
                        stats["unchanged"] += 1

                else:
                    # New listing
                    slug = generate_unique_slug(listing.get("title", ""), existing_slugs)
                    logger.info(
                        f"NEW {cpa_key}: {listing.get('title', 'N/A')} | "
                        f"{listing.get('property_type', '?')} | "
                        f"{listing.get('listing_type', '?')} | "
                        f"price={listing.get('price', '?')}"
                    )

                    if not dry_run:
                        sql = build_insert_sql(listing, slug)
                        sql_batch.append(sql)

                    stats["new"] += 1
                    stats["new_cpa_codes"].append(cpa_key)

            except Exception as e:
                stats["errors"] += 1
                logger.error(f"Error processing {cpa_key} (sheet {listing.get('_sheet')}, row {listing.get('_row')}): {e}")
                continue

            # Execute batch when it reaches the batch size
            if not dry_run and len(sql_batch) >= SQL_BATCH_SIZE:
                executed_batch_count += 1
                logger.info(f"Executing SQL batch #{executed_batch_count} ({len(sql_batch)} statements)...")
                if not run_sql_batch(sql_batch, logger):
                    logger.error(f"Batch #{executed_batch_count} failed! Some updates may not have been applied.")
                    # Don't count all statements as errors, just the batch
                    stats["errors"] += 1
                    # Clear batch to prevent cascade errors
                    sql_batch = []
                else:
                    sql_batch = []

            # Progress logging
            if (idx + 1) % 5000 == 0:
                logger.info(
                    f"Progress: {idx + 1}/{len(all_sheet_data)} | "
                    f"New={stats['new']} Updated={stats['updated']} Unchanged={stats['unchanged']}"
                )

        # Execute remaining batch
        if not dry_run and sql_batch:
            executed_batch_count += 1
            logger.info(f"Executing final SQL batch #{executed_batch_count} ({len(sql_batch)} statements)...")
            if not run_sql_batch(sql_batch, logger):
                logger.error(f"Final batch failed!")
                stats["errors"] += 1

        # 5. Trigger image sync for new listings if requested
        if sync_images and stats["new"] > 0 and not dry_run:
            logger.info(f"Triggering image sync for {stats['new']} new listings...")
            try:
                result = subprocess.run(
                    ["python3", SYNC_IMAGES_SCRIPT, str(stats["new"] + 10)],
                    capture_output=True, text=True, timeout=3600
                )
                logger.info(f"Image sync completed (exit code: {result.returncode})")
                if result.stdout:
                    for line in result.stdout.strip().split("\n")[-5:]:
                        logger.info(f"  [images] {line}")
            except Exception as e:
                logger.error(f"Image sync failed: {e}")

    except Exception as e:
        logger.error(f"FATAL ERROR: {e}")
        import traceback
        logger.error(traceback.format_exc())
        stats["errors"] += 1

    # 6. Print summary
    logger.info("")
    logger.info("=" * 70)
    logger.info("SYNC SUMMARY")
    logger.info("=" * 70)
    logger.info(f"Sheet 1 rows processed: {stats['sheet1_rows']}")
    logger.info(f"Sheet 2 rows processed: {stats['sheet2_rows']}")
    logger.info(f"New listings inserted:  {stats['new']}")
    logger.info(f"Listings updated:       {stats['updated']}")
    logger.info(f"Listings deactivated:   {stats['deactivated']}")
    logger.info(f"Unchanged:              {stats['unchanged']}")
    logger.info(f"Skipped (no title):     {stats['skipped']}")
    logger.info(f"Errors:                 {stats['errors']}")
    if executed_batch_count > 0:
        logger.info(f"SQL batches executed:   {executed_batch_count}")
    logger.info("=" * 70)

    if stats["new_cpa_codes"] and stats["new"] <= 50:
        logger.info(f"New CPA codes: {', '.join(stats['new_cpa_codes'])}")

    logger.info(f"SYNC COMPLETED at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    return stats


# ─── Entry Point ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Sync listings from Google Sheets to PostgreSQL")
    parser.add_argument("--dry-run", action="store_true", help="Preview changes without writing to DB")
    parser.add_argument("--sheet", type=int, default=0, choices=[0, 1, 2],
                        help="Which sheet to sync (0=both, 1=JUAL&SEWA, 2=DEVELOPER)")
    parser.add_argument("--sync-images", action="store_true", help="Trigger image sync after listing sync")

    args = parser.parse_args()
    stats = sync(args)
    sys.exit(1 if stats["errors"] > 0 else 0)
